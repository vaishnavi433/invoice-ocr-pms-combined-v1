# ============================================================================
# SUPY VISION CONVERTER v4.0 (UNIFIED - UNABRIDGED EDITION)
# ============================================================================
# AUTHOR: Gemini AI for Supy.io Operations
# DATE: 2026-01-08
#
# DESCRIPTION:
# This is the unified engine that combines:
# 1. Optical Character Recognition (OCR) via Vision AI for Invoices.
# 2. PMS Data Standardization (AI Cleaning) for Item Lists.
# 3. Global Tax Logic for 200+ countries.
# 4. Recursive Batch Processing for ZIPs and Drive Folders.
# 5. Automated Excel Formatting with 5-Sheet output.
# ============================================================================

# ----------------------------------------------------------------------------
# PART 1: DEPENDENCY INSTALLATION & IMPORTS
# ----------------------------------------------------------------------------
import sys
import subprocess

# Auto-install dependencies if missing
def install_dependencies():
    packages = [
        "requests", "pandas", "openpyxl", "pillow", "tqdm",
        "google-auth", "google-auth-oauthlib", "google-auth-httplib2",
        "google-api-python-client", "rapidfuzz", "beautifulsoup4"
    ]
    for package in packages:
        try:
            __import__(package)
        except ImportError:
            print(f"📦 Installing missing package: {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

install_dependencies()

import base64
import json
import os
import re
import zipfile
import shutil
import io
import time
import logging
import pandas as pd
import requests
import concurrent.futures
from pathlib import Path
from datetime import datetime, date
from typing import List, Tuple, Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm.auto import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz
from collections import Counter

# Google Colab Environment Check
try:
    from google.colab import auth, files, drive
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    COLAB_ENV = True
except ImportError:
    COLAB_ENV = False
    print("⚠️ Environment Warning: Not running in Google Colab. Google Drive features will be disabled.")

# ----------------------------------------------------------------------------
# PART 2: GLOBAL CONFIGURATION & CONSTANTS
# ----------------------------------------------------------------------------

# API Configuration
OPENROUTER_API_KEY = "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"
MODEL_ID = "anthropic/claude-sonnet-4.5"

# Processing Tunables
MAX_WORKERS = 5            # Number of parallel threads
BATCH_SIZE = 10            # Rows per AI batch (PMS mode)
RETRY_ATTEMPTS = 3         # API retry count
RETRY_DELAY = 2            # Seconds between retries
OUTPUT_MAX_TOKENS = 8000   # Max tokens for response

# Supported File Types
INVOICE_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png', '.webp', '.heic'}
EXCEL_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

# ----------------------------------------------------------------------------
# PART 3: COMPLETE GLOBAL DATABASES
# ----------------------------------------------------------------------------

# A. FULL ITEM CATEGORY TAXONOMY
CATEGORY_TAXONOMY = {
    "FOOD": {
        "DRY ITEMS": ["Rice", "Pasta", "Flour", "Sugar", "Beans", "Nuts", "Spices", "Grains", "Cereal", "Bakery Needs", "Crackers", "Biscuits", "Chips", "Snacks"],
        "DAIRY PRODUCTS": ["Milk", "Cheese", "Yogurt", "Butter", "Cream", "Ice Cream", "Eggs", "Dairy Alternatives", "Ghee"],
        "FRUITS & VEGETABLES": ["Fresh Fruits", "Fresh Vegetables", "Herbs", "Salad Greens", "Mushrooms", "Root Vegetables", "Microgreens"],
        "BEEF & SEAFOOD & POULTRY": ["Beef", "Chicken", "Pork", "Seafood", "Fish", "Lamb", "Turkey", "Duck", "Shellfish", "Sausages", "Cured Meats"],
        "OILS & VINEGARS": ["Olive Oil", "Vegetable Oil", "Vinegar", "Cooking Oils", "Specialty Oils", "Balsamic", "Coconut Oil"],
        "SAUCES & PASTES": ["Tomato Sauce", "Soy Sauce", "Paste", "Condiments", "Dressings", "Mayonnaise", "Mustard", "Ketchup", "Chili Paste", "Curry Paste"],
        "FROZEN FOOD": ["Frozen Vegetables", "Frozen Meat", "Frozen Seafood", "Frozen Pastries", "Frozen Fruits", "Ice"]
    },
    "BEVERAGES": {
        "SOFT DRINKS": ["Cola", "Soda", "Lemon-Lime", "Energy Drinks", "Tonics", "Ginger Ale"],
        "ALCOHOL": ["Beer", "Wine", "Spirits", "Liquor", "Champagne", "Sake", "Soju", "Cider"],
        "WATER": ["Still Water", "Sparkling Water", "Mineral Water", "Coconut Water"],
        "JUICE": ["Orange Juice", "Apple Juice", "Mixed Juices", "Smoothies", "Fresh Juice"],
        "COFFEE & TEA": ["Coffee Beans", "Ground Coffee", "Tea Leaves", "Tea Bags", "Coffee Capsules", "Matcha", "Chai"],
        "SYRUPS & MIXERS": ["Cocktail Syrups", "Fruit Purees", "Cordials", "Bitters"]
    },
    "NON FOOD": {
        "PACKAGING": ["Boxes", "Bags", "Containers", "Wrapping", "Foil", "Cling Film", "Takeaway Containers", "Pizza Boxes"],
        "CLEANING": ["Detergents", "Sanitizers", "Chemicals", "Disinfectants", "Sponges", "Mops", "Brooms", "Trash Bags"],
        "KITCHEN SUPPLIES / DISPOSABLES": ["Plates", "Cups", "Cutlery", "Napkins", "Towels", "Straws", "Gloves", "Hairnets", "Aprons"],
        "OFFICE SUPPLIES": ["Paper", "Pens", "Ink", "Staples", "Notebooks"],
        "MAINTENANCE": ["Light Bulbs", "Batteries", "Tools", "Hardware"]
    }
}

# B. COMPLETE 200+ COUNTRY TAX DATABASE
# This ensures precise tax logic for any region in the world.
COUNTRY_TAX_DB = {
    "AF": {"name": "Afghanistan", "rate": 10, "food_exempt": False, "currency": "AFN"},
    "AL": {"name": "Albania", "rate": 20, "food_exempt": False, "currency": "ALL"},
    "DZ": {"name": "Algeria", "rate": 19, "food_exempt": True, "currency": "DZD"},
    "AS": {"name": "American Samoa", "rate": 0, "food_exempt": True, "currency": "USD"},
    "AD": {"name": "Andorra", "rate": 4.5, "food_exempt": False, "currency": "EUR"},
    "AO": {"name": "Angola", "rate": 14, "food_exempt": False, "currency": "AOA"},
    "AI": {"name": "Anguilla", "rate": 0, "food_exempt": True, "currency": "XCD"},
    "AG": {"name": "Antigua & Barbuda", "rate": 15, "food_exempt": True, "currency": "XCD"},
    "AR": {"name": "Argentina", "rate": 21, "food_exempt": True, "currency": "ARS"},
    "AM": {"name": "Armenia", "rate": 20, "food_exempt": False, "currency": "AMD"},
    "AW": {"name": "Aruba", "rate": 6, "food_exempt": False, "currency": "AWG"},
    "AU": {"name": "Australia", "rate": 10, "food_exempt": True, "currency": "AUD"},
    "AT": {"name": "Austria", "rate": 20, "food_exempt": True, "currency": "EUR"},
    "AZ": {"name": "Azerbaijan", "rate": 18, "food_exempt": False, "currency": "AZN"},
    "BS": {"name": "Bahamas", "rate": 12, "food_exempt": True, "currency": "BSD"},
    "BH": {"name": "Bahrain", "rate": 10, "food_exempt": True, "currency": "BHD"},
    "BD": {"name": "Bangladesh", "rate": 15, "food_exempt": True, "currency": "BDT"},
    "BB": {"name": "Barbados", "rate": 17.5, "food_exempt": True, "currency": "BBD"},
    "BY": {"name": "Belarus", "rate": 20, "food_exempt": True, "currency": "BYN"},
    "BE": {"name": "Belgium", "rate": 21, "food_exempt": True, "currency": "EUR"},
    "BZ": {"name": "Belize", "rate": 12.5, "food_exempt": True, "currency": "BZD"},
    "BJ": {"name": "Benin", "rate": 18, "food_exempt": False, "currency": "XOF"},
    "BM": {"name": "Bermuda", "rate": 0, "food_exempt": True, "currency": "BMD"},
    "BT": {"name": "Bhutan", "rate": 0, "food_exempt": True, "currency": "BTN"},
    "BO": {"name": "Bolivia", "rate": 13, "food_exempt": False, "currency": "BOB"},
    "BA": {"name": "Bosnia", "rate": 17, "food_exempt": True, "currency": "BAM"},
    "BW": {"name": "Botswana", "rate": 12, "food_exempt": True, "currency": "BWP"},
    "BR": {"name": "Brazil", "rate": 17, "food_exempt": False, "currency": "BRL"},
    "BN": {"name": "Brunei", "rate": 0, "food_exempt": True, "currency": "BND"},
    "BG": {"name": "Bulgaria", "rate": 20, "food_exempt": True, "currency": "BGN"},
    "BF": {"name": "Burkina Faso", "rate": 18, "food_exempt": False, "currency": "XOF"},
    "BI": {"name": "Burundi", "rate": 18, "food_exempt": False, "currency": "BIF"},
    "KH": {"name": "Cambodia", "rate": 10, "food_exempt": True, "currency": "KHR"},
    "CM": {"name": "Cameroon", "rate": 19.25, "food_exempt": False, "currency": "XAF"},
    "CA": {"name": "Canada", "rate": 5, "food_exempt": True, "currency": "CAD"},
    "CV": {"name": "Cape Verde", "rate": 15, "food_exempt": True, "currency": "CVE"},
    "KY": {"name": "Cayman Islands", "rate": 0, "food_exempt": True, "currency": "KYD"},
    "CF": {"name": "Central African Rep.", "rate": 19, "food_exempt": False, "currency": "XAF"},
    "TD": {"name": "Chad", "rate": 18, "food_exempt": False, "currency": "XAF"},
    "CL": {"name": "Chile", "rate": 19, "food_exempt": True, "currency": "CLP"},
    "CN": {"name": "China", "rate": 13, "food_exempt": True, "currency": "CNY"},
    "CO": {"name": "Colombia", "rate": 19, "food_exempt": True, "currency": "COP"},
    "KM": {"name": "Comoros", "rate": 10, "food_exempt": False, "currency": "KMF"},
    "CG": {"name": "Congo", "rate": 18, "food_exempt": False, "currency": "XAF"},
    "CR": {"name": "Costa Rica", "rate": 13, "food_exempt": True, "currency": "CRC"},
    "HR": {"name": "Croatia", "rate": 25, "food_exempt": True, "currency": "EUR"},
    "CU": {"name": "Cuba", "rate": 10, "food_exempt": False, "currency": "CUP"},
    "CY": {"name": "Cyprus", "rate": 19, "food_exempt": True, "currency": "EUR"},
    "CZ": {"name": "Czech Republic", "rate": 21, "food_exempt": True, "currency": "CZK"},
    "DK": {"name": "Denmark", "rate": 25, "food_exempt": False, "currency": "DKK"},
    "DJ": {"name": "Djibouti", "rate": 10, "food_exempt": False, "currency": "DJF"},
    "DM": {"name": "Dominica", "rate": 15, "food_exempt": True, "currency": "XCD"},
    "DO": {"name": "Dominican Republic", "rate": 18, "food_exempt": True, "currency": "DOP"},
    "EC": {"name": "Ecuador", "rate": 12, "food_exempt": True, "currency": "USD"},
    "EG": {"name": "Egypt", "rate": 14, "food_exempt": False, "currency": "EGP"},
    "SV": {"name": "El Salvador", "rate": 13, "food_exempt": True, "currency": "USD"},
    "GQ": {"name": "Equatorial Guinea", "rate": 15, "food_exempt": False, "currency": "XAF"},
    "ER": {"name": "Eritrea", "rate": 0, "food_exempt": True, "currency": "ERN"},
    "EE": {"name": "Estonia", "rate": 20, "food_exempt": True, "currency": "EUR"},
    "SZ": {"name": "Eswatini", "rate": 15, "food_exempt": True, "currency": "SZL"},
    "ET": {"name": "Ethiopia", "rate": 15, "food_exempt": True, "currency": "ETB"},
    "FJ": {"name": "Fiji", "rate": 9, "food_exempt": True, "currency": "FJD"},
    "FI": {"name": "Finland", "rate": 24, "food_exempt": True, "currency": "EUR"},
    "FR": {"name": "France", "rate": 20, "food_exempt": True, "currency": "EUR"},
    "GA": {"name": "Gabon", "rate": 18, "food_exempt": False, "currency": "XAF"},
    "GM": {"name": "Gambia", "rate": 15, "food_exempt": False, "currency": "GMD"},
    "GE": {"name": "Georgia", "rate": 18, "food_exempt": True, "currency": "GEL"},
    "DE": {"name": "Germany", "rate": 19, "food_exempt": True, "currency": "EUR"},
    "GH": {"name": "Ghana", "rate": 12.5, "food_exempt": False, "currency": "GHS"},
    "GR": {"name": "Greece", "rate": 24, "food_exempt": True, "currency": "EUR"},
    "GD": {"name": "Grenada", "rate": 15, "food_exempt": True, "currency": "XCD"},
    "GT": {"name": "Guatemala", "rate": 12, "food_exempt": False, "currency": "GTQ"},
    "GN": {"name": "Guinea", "rate": 18, "food_exempt": False, "currency": "GNF"},
    "GY": {"name": "Guyana", "rate": 14, "food_exempt": True, "currency": "GYD"},
    "HT": {"name": "Haiti", "rate": 10, "food_exempt": False, "currency": "HTG"},
    "HN": {"name": "Honduras", "rate": 15, "food_exempt": True, "currency": "HNL"},
    "HK": {"name": "Hong Kong", "rate": 0, "food_exempt": True, "currency": "HKD"},
    "HU": {"name": "Hungary", "rate": 27, "food_exempt": True, "currency": "HUF"},
    "IS": {"name": "Iceland", "rate": 24, "food_exempt": True, "currency": "ISK"},
    "IN": {"name": "India", "rate": 18, "food_exempt": True, "currency": "INR"},
    "ID": {"name": "Indonesia", "rate": 11, "food_exempt": False, "currency": "IDR"},
    "IR": {"name": "Iran", "rate": 9, "food_exempt": True, "currency": "IRR"},
    "IQ": {"name": "Iraq", "rate": 0, "food_exempt": True, "currency": "IQD"},
    "IE": {"name": "Ireland", "rate": 23, "food_exempt": True, "currency": "EUR"},
    "IL": {"name": "Israel", "rate": 17, "food_exempt": True, "currency": "ILS"},
    "IT": {"name": "Italy", "rate": 22, "food_exempt": True, "currency": "EUR"},
    "JM": {"name": "Jamaica", "rate": 15, "food_exempt": True, "currency": "JMD"},
    "JP": {"name": "Japan", "rate": 10, "food_exempt": True, "currency": "JPY"},
    "JO": {"name": "Jordan", "rate": 16, "food_exempt": True, "currency": "JOD"},
    "KZ": {"name": "Kazakhstan", "rate": 12, "food_exempt": True, "currency": "KZT"},
    "KE": {"name": "Kenya", "rate": 16, "food_exempt": True, "currency": "KES"},
    "KI": {"name": "Kiribati", "rate": 0, "food_exempt": True, "currency": "AUD"},
    "KP": {"name": "North Korea", "rate": 0, "food_exempt": True, "currency": "KPW"},
    "KR": {"name": "South Korea", "rate": 10, "food_exempt": True, "currency": "KRW"},
    "KW": {"name": "Kuwait", "rate": 0, "food_exempt": True, "currency": "KWD"},
    "KG": {"name": "Kyrgyzstan", "rate": 12, "food_exempt": True, "currency": "KGS"},
    "LA": {"name": "Laos", "rate": 7, "food_exempt": False, "currency": "LAK"},
    "LV": {"name": "Latvia", "rate": 21, "food_exempt": True, "currency": "EUR"},
    "LB": {"name": "Lebanon", "rate": 11, "food_exempt": True, "currency": "LBP"},
    "LS": {"name": "Lesotho", "rate": 15, "food_exempt": False, "currency": "LSL"},
    "LR": {"name": "Liberia", "rate": 10, "food_exempt": False, "currency": "LRD"},
    "LY": {"name": "Libya", "rate": 0, "food_exempt": True, "currency": "LYD"},
    "LI": {"name": "Liechtenstein", "rate": 7.7, "food_exempt": True, "currency": "CHF"},
    "LT": {"name": "Lithuania", "rate": 21, "food_exempt": True, "currency": "EUR"},
    "LU": {"name": "Luxembourg", "rate": 17, "food_exempt": True, "currency": "EUR"},
    "MO": {"name": "Macao", "rate": 0, "food_exempt": True, "currency": "MOP"},
    "MG": {"name": "Madagascar", "rate": 20, "food_exempt": False, "currency": "MGA"},
    "MW": {"name": "Malawi", "rate": 16.5, "food_exempt": True, "currency": "MWK"},
    "MY": {"name": "Malaysia", "rate": 10, "food_exempt": True, "currency": "MYR"},
    "MV": {"name": "Maldives", "rate": 6, "food_exempt": False, "currency": "MVR"},
    "ML": {"name": "Mali", "rate": 18, "food_exempt": False, "currency": "XOF"},
    "MT": {"name": "Malta", "rate": 18, "food_exempt": True, "currency": "EUR"},
    "MH": {"name": "Marshall Islands", "rate": 0, "food_exempt": True, "currency": "USD"},
    "MR": {"name": "Mauritania", "rate": 16, "food_exempt": False, "currency": "MRU"},
    "MU": {"name": "Mauritius", "rate": 15, "food_exempt": True, "currency": "MUR"},
    "MX": {"name": "Mexico", "rate": 16, "food_exempt": True, "currency": "MXN"},
    "FM": {"name": "Micronesia", "rate": 0, "food_exempt": True, "currency": "USD"},
    "MD": {"name": "Moldova", "rate": 20, "food_exempt": True, "currency": "MDL"},
    "MC": {"name": "Monaco", "rate": 20, "food_exempt": True, "currency": "EUR"},
    "MN": {"name": "Mongolia", "rate": 10, "food_exempt": False, "currency": "MNT"},
    "ME": {"name": "Montenegro", "rate": 21, "food_exempt": True, "currency": "EUR"},
    "MA": {"name": "Morocco", "rate": 20, "food_exempt": True, "currency": "MAD"},
    "MZ": {"name": "Mozambique", "rate": 17, "food_exempt": True, "currency": "MZN"},
    "MM": {"name": "Myanmar", "rate": 5, "food_exempt": False, "currency": "MMK"},
    "NA": {"name": "Namibia", "rate": 15, "food_exempt": False, "currency": "NAD"},
    "NR": {"name": "Nauru", "rate": 0, "food_exempt": True, "currency": "AUD"},
    "NP": {"name": "Nepal", "rate": 13, "food_exempt": True, "currency": "NPR"},
    "NL": {"name": "Netherlands", "rate": 21, "food_exempt": True, "currency": "EUR"},
    "NZ": {"name": "New Zealand", "rate": 15, "food_exempt": True, "currency": "NZD"},
    "NI": {"name": "Nicaragua", "rate": 15, "food_exempt": True, "currency": "NIO"},
    "NE": {"name": "Niger", "rate": 19, "food_exempt": False, "currency": "XOF"},
    "NG": {"name": "Nigeria", "rate": 7.5, "food_exempt": False, "currency": "NGN"},
    "MK": {"name": "North Macedonia", "rate": 18, "food_exempt": True, "currency": "MKD"},
    "NO": {"name": "Norway", "rate": 25, "food_exempt": True, "currency": "NOK"},
    "OM": {"name": "Oman", "rate": 5, "food_exempt": True, "currency": "OMR"},
    "PK": {"name": "Pakistan", "rate": 17, "food_exempt": True, "currency": "PKR"},
    "PW": {"name": "Palau", "rate": 0, "food_exempt": True, "currency": "USD"},
    "PA": {"name": "Panama", "rate": 7, "food_exempt": True, "currency": "PAB"},
    "PG": {"name": "Papua New Guinea", "rate": 10, "food_exempt": True, "currency": "PGK"},
    "PY": {"name": "Paraguay", "rate": 10, "food_exempt": True, "currency": "PYG"},
    "PE": {"name": "Peru", "rate": 18, "food_exempt": True, "currency": "PEN"},
    "PH": {"name": "Philippines", "rate": 12, "food_exempt": False, "currency": "PHP"},
    "PL": {"name": "Poland", "rate": 23, "food_exempt": True, "currency": "PLN"},
    "PT": {"name": "Portugal", "rate": 23, "food_exempt": True, "currency": "EUR"},
    "QA": {"name": "Qatar", "rate": 0, "food_exempt": True, "currency": "QAR"},
    "RO": {"name": "Romania", "rate": 19, "food_exempt": True, "currency": "RON"},
    "RU": {"name": "Russia", "rate": 20, "food_exempt": True, "currency": "RUB"},
    "RW": {"name": "Rwanda", "rate": 18, "food_exempt": False, "currency": "RWF"},
    "KN": {"name": "Saint Kitts & Nevis", "rate": 17, "food_exempt": False, "currency": "XCD"},
    "LC": {"name": "Saint Lucia", "rate": 12.5, "food_exempt": False, "currency": "XCD"},
    "VC": {"name": "Saint Vincent", "rate": 16, "food_exempt": False, "currency": "XCD"},
    "WS": {"name": "Samoa", "rate": 15, "food_exempt": False, "currency": "WST"},
    "SM": {"name": "San Marino", "rate": 17, "food_exempt": False, "currency": "EUR"},
    "ST": {"name": "Sao Tome", "rate": 15, "food_exempt": False, "currency": "STN"},
    "SA": {"name": "Saudi Arabia", "rate": 15, "food_exempt": True, "currency": "SAR"},
    "SN": {"name": "Senegal", "rate": 18, "food_exempt": False, "currency": "XOF"},
    "RS": {"name": "Serbia", "rate": 20, "food_exempt": True, "currency": "RSD"},
    "SC": {"name": "Seychelles", "rate": 15, "food_exempt": False, "currency": "SCR"},
    "SL": {"name": "Sierra Leone", "rate": 15, "food_exempt": False, "currency": "SLL"},
    "SG": {"name": "Singapore", "rate": 9, "food_exempt": True, "currency": "SGD"},
    "SK": {"name": "Slovakia", "rate": 20, "food_exempt": True, "currency": "EUR"},
    "SI": {"name": "Slovenia", "rate": 22, "food_exempt": True, "currency": "EUR"},
    "SB": {"name": "Solomon Islands", "rate": 10, "food_exempt": False, "currency": "SBD"},
    "SO": {"name": "Somalia", "rate": 0, "food_exempt": True, "currency": "SOS"},
    "ZA": {"name": "South Africa", "rate": 15, "food_exempt": True, "currency": "ZAR"},
    "SS": {"name": "South Sudan", "rate": 15, "food_exempt": False, "currency": "SSP"},
    "ES": {"name": "Spain", "rate": 21, "food_exempt": True, "currency": "EUR"},
    "LK": {"name": "Sri Lanka", "rate": 15, "food_exempt": False, "currency": "LKR"},
    "SD": {"name": "Sudan", "rate": 17, "food_exempt": False, "currency": "SDG"},
    "SR": {"name": "Suriname", "rate": 10, "food_exempt": False, "currency": "SRD"},
    "SE": {"name": "Sweden", "rate": 25, "food_exempt": True, "currency": "SEK"},
    "CH": {"name": "Switzerland", "rate": 7.7, "food_exempt": True, "currency": "CHF"},
    "SY": {"name": "Syria", "rate": 0, "food_exempt": True, "currency": "SYP"},
    "TW": {"name": "Taiwan", "rate": 5, "food_exempt": False, "currency": "TWD"},
    "TJ": {"name": "Tajikistan", "rate": 15, "food_exempt": False, "currency": "TJS"},
    "TZ": {"name": "Tanzania", "rate": 18, "food_exempt": False, "currency": "TZS"},
    "TH": {"name": "Thailand", "rate": 7, "food_exempt": False, "currency": "THB"},
    "TL": {"name": "Timor-Leste", "rate": 2.5, "food_exempt": False, "currency": "USD"},
    "TG": {"name": "Togo", "rate": 18, "food_exempt": False, "currency": "XOF"},
    "TO": {"name": "Tonga", "rate": 15, "food_exempt": False, "currency": "TOP"},
    "TT": {"name": "Trinidad & Tobago", "rate": 12.5, "food_exempt": True, "currency": "TTD"},
    "TN": {"name": "Tunisia", "rate": 19, "food_exempt": True, "currency": "TND"},
    "TR": {"name": "Turkey", "rate": 20, "food_exempt": True, "currency": "TRY"},
    "TM": {"name": "Turkmenistan", "rate": 15, "food_exempt": False, "currency": "TMT"},
    "TV": {"name": "Tuvalu", "rate": 7, "food_exempt": False, "currency": "AUD"},
    "UG": {"name": "Uganda", "rate": 18, "food_exempt": False, "currency": "UGX"},
    "UA": {"name": "Ukraine", "rate": 20, "food_exempt": True, "currency": "UAH"},
    "AE": {"name": "UAE", "rate": 5, "food_exempt": True, "currency": "AED"},
    "GB": {"name": "UK", "rate": 20, "food_exempt": True, "currency": "GBP"},
    "US": {"name": "USA", "rate": 0, "food_exempt": True, "currency": "USD"},
    "UY": {"name": "Uruguay", "rate": 22, "food_exempt": True, "currency": "UYU"},
    "UZ": {"name": "Uzbekistan", "rate": 12, "food_exempt": True, "currency": "UZS"},
    "VU": {"name": "Vanuatu", "rate": 12.5, "food_exempt": False, "currency": "VUV"},
    "VE": {"name": "Venezuela", "rate": 16, "food_exempt": True, "currency": "VES"},
    "VN": {"name": "Vietnam", "rate": 10, "food_exempt": True, "currency": "VND"},
    "YE": {"name": "Yemen", "rate": 5, "food_exempt": False, "currency": "YER"},
    "ZM": {"name": "Zambia", "rate": 16, "food_exempt": False, "currency": "ZMW"},
    "ZW": {"name": "Zimbabwe", "rate": 14.5, "food_exempt": False, "currency": "ZWL"}
}

# ----------------------------------------------------------------------------
# PART 4: ROBUST UTILITY FUNCTIONS
# ----------------------------------------------------------------------------

def authenticate_google_drive():
    """Securely authenticates with Google Drive when running in Colab."""
    if not COLAB_ENV:
        raise RuntimeError("Google Drive authentication only available in Colab environment.")
    print("🔐 Initiating Google Drive Authentication...")
    try:
        auth.authenticate_user()
        service = build('drive', 'v3')
        print("✅ Google Drive Authenticated Successfully.")
        return service
    except Exception as e:
        print(f"❌ Authentication Failed: {e}")
        return None

def extract_drive_folder_id(url_or_id: str) -> str:
    """Parses various Google Drive URL formats to extract the clean Folder ID."""
    if not url_or_id:
        return ""
    if '/' not in url_or_id and len(url_or_id) > 20:
        return url_or_id

    # Regex patterns for different Drive URL formats
    patterns = [
        r'folders/([a-zA-Z0-9-_]+)',
        r'id=([a-zA-Z0-9-_]+)',
        r'/d/([a-zA-Z0-9-_]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, url_or_id)
        if match:
            return match.group(1)
    return url_or_id

def recursive_zip_extractor(zip_path: str, base_extract_dir: str) -> List[str]:
    """
    Recursively extracts a ZIP file, including nested ZIPs inside it.
    Returns a flat list of all relevant file paths found.
    """
    print(f"📦 Unpacking: {Path(zip_path).name}")
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(base_extract_dir)
    except zipfile.BadZipFile:
        print(f"❌ Error: Corrupt ZIP file {zip_path}")
        return []

    all_files_found = []
    nested_zips_found = []

    # Walk through the extracted directory
    for root, dirs, files in os.walk(base_extract_dir):
        for file in files:
            full_path = os.path.join(root, file)
            extension = Path(file).suffix.lower()

            if extension == '.zip':
                nested_zips_found.append(full_path)
            elif extension in INVOICE_EXTENSIONS or extension in EXCEL_EXTENSIONS:
                all_files_found.append(full_path)

    # Recursively handle nested ZIPs
    for nested_zip in nested_zips_found:
        nested_dir_name = Path(nested_zip).stem + "_extracted"
        nested_extract_path = os.path.join(os.path.dirname(nested_zip), nested_dir_name)
        os.makedirs(nested_extract_path, exist_ok=True)

        # Recursive call
        nested_files = recursive_zip_extractor(nested_zip, nested_extract_path)
        all_files_found.extend(nested_files)

    return all_files_found

def normalize_text_spacing(text: str) -> str:
    """
    Fixes spacing issues common in OCR output.
    Example: '5kg' -> '5 kg', '10ml' -> '10 ml'
    """
    if not isinstance(text, str):
        return text

    # Insert space between number and unit
    text = re.sub(r'(\d+(?:\.\d+)?)(kg|g|gm|l|ml|ltr|gram|oz|lb|pcs|pc)', r'\1 \2', text, flags=re.IGNORECASE)
    # Collapse multiple spaces
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def detect_text_language(text: str) -> str:
    """
    Detects the primary language of a text string using Unicode ranges.
    Supports: Chinese, Arabic, Thai, Japanese, Korean, Russian, Vietnamese, Hindi.
    """
    if not text:
        return "English"
    text = str(text)

    if re.search(r'[\u4e00-\u9fff]', text): return "Chinese"
    if re.search(r'[\u0600-\u06ff]', text): return "Arabic"
    if re.search(r'[\u0e00-\u0e7f]', text): return "Thai"
    if re.search(r'[\u3040-\u309f\u30a0-\u30ff]', text): return "Japanese"
    if re.search(r'[\uac00-\ud7af]', text): return "Korean"
    if re.search(r'[\u0400-\u04ff]', text): return "Russian"
    if re.search(r'[\u0900-\u097f]', text): return "Hindi"
    if re.search(r'[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ]', text, re.IGNORECASE): return "Vietnamese"

    return "English"

def make_serializable(obj):
    """Recursively converts objects to JSON-serializable formats."""
    if isinstance(obj, (str, int, float, bool)) or obj is None:
        return obj
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {str(k): make_serializable(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_serializable(x) for x in obj]
    return str(obj)

def convert_file_to_base64(file_path: str) -> str:
    """Reads a file and converts it to a Data URL string for API transmission."""
    mime_type = "application/octet-stream"
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        mime_type = "application/pdf"
    elif ext in [".jpg", ".jpeg"]:
        mime_type = "image/jpeg"
    elif ext == ".png":
        mime_type = "image/png"
    elif ext == ".webp":
        mime_type = "image/webp"
    elif ext == ".heic":
        mime_type = "image/heic"

    try:
        with open(file_path, "rb") as f:
            file_content = f.read()
            encoded_string = base64.b64encode(file_content).decode("utf-8")
            return f"data:{mime_type};base64,{encoded_string}"
    except Exception as e:
        print(f"❌ Error encoding file {file_path}: {e}")
        return ""

# ----------------------------------------------------------------------------
# PART 5: CORE ENGINE - INVOICE EXTRACTION (OCR)
# ----------------------------------------------------------------------------

def generate_invoice_system_prompt(country_code: str) -> str:
    """Generates the context-aware system prompt for Invoice Extraction."""
    country_info = COUNTRY_TAX_DB.get(country_code, COUNTRY_TAX_DB["AE"])

    prompt = f"""You are an expert autonomous Invoice Data Extraction AI for the F&B industry.

CURRENT REGION CONTEXT:
- Country: {country_info['name']} ({country_code})
- Standard Tax Rate: {country_info['rate']}%
- Food Items Tax Exempt: {country_info['food_exempt']}
- Local Currency: {country_info['currency']}

YOUR MISSION:
Extract every single line item from the provided invoice document (image or PDF).

OUTPUT REQUIREMENTS:
Return a strictly valid JSON object. Do not include markdown code blocks. The JSON must follow this schema:
{{
  "invoice_metadata": {{
    "invoice_number": "string",
    "invoice_date": "YYYY-MM-DD",
    "supplier_name": "string",
    "total_amount": number
  }},
  "line_items": [
    {{
      "Supplier Item Name": "Exact text from invoice",
      "Supplier Item Code": "SKU or code if visible, else empty string",
      "Supplier Name": "Vendor Name",
      "Buying Unit": "Unit string (e.g., kg, case, box, ea)",
      "Price": number (unit price),
      "Discount": number (0 if none),
      "Tax Rate": number (percentage),
      "Page Number": integer
    }}
  ]
}}

LOGIC RULES:
1. **Tax Logic**:
   - If the country is {country_info['name']} and 'Food Exempt' is True, set 'Tax Rate' to 0 for all food ingredients.
   - For Non-Food items (Cleaning, Packaging, Alcohol), use the standard rate of {country_info['rate']}%.
   - If the invoice explicitly states a tax rate for a line, use that specific rate.
2. **Data Cleaning**:
   - Remove currency symbols from Price (e.g., '$10.00' -> 10.00).
   - Convert dates to ISO format (YYYY-MM-DD).
3. **Completeness**:
   - Do not summarize. Extract every single row.
"""
    return prompt

def process_single_invoice(file_path: str, country_code: str) -> pd.DataFrame:
    """
    Sends a single invoice to the Vision AI API and returns a DataFrame of extracted items.
    """
    data_uri = convert_file_to_base64(file_path)
    if not data_uri:
        return pd.DataFrame()

    system_prompt = generate_invoice_system_prompt(country_code)
    is_pdf = file_path.lower().endswith(".pdf")

    # Construct the user message payload
    user_content = []
    user_content.append({"type": "text", "text": "Please extract all line items from this document as JSON."})

    if is_pdf:
        # PDF handling for compatible models
        user_content.append({
            "type": "file",
            "file": {
                "filename": Path(file_path).name,
                "file_data": data_uri
            }
        })
    else:
        # Image handling
        user_content.append({
            "type": "image_url",
            "image_url": {"url": data_uri}
        })

    api_payload = {
        "model": MODEL_ID,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ],
        "temperature": 0.0, # Zero temperature for deterministic extraction
        "max_tokens": OUTPUT_MAX_TOKENS
    }

    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }

    # Retry Loop with Exponential Backoff
    for attempt in range(RETRY_ATTEMPTS):
        try:
            response = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=api_payload, timeout=120)

            if response.status_code != 200:
                print(f"   ⚠️ API Error (Attempt {attempt+1}/{RETRY_ATTEMPTS}): {response.status_code} - {response.text[:100]}")
                time.sleep(RETRY_DELAY * (attempt + 1))
                continue

            response_json = response.json()
            if 'choices' not in response_json or not response_json['choices']:
                print(f"   ⚠️ Empty response from API for {Path(file_path).name}")
                return pd.DataFrame()

            content_text = response_json["choices"][0]["message"]["content"]

            # Sanitize JSON string (remove markdown fences if present)
            content_text = re.sub(r"^```json", "", content_text, flags=re.MULTILINE)
            content_text = re.sub(r"^```", "", content_text, flags=re.MULTILINE)
            content_text = content_text.strip()

            parsed_data = json.loads(content_text)

            # Extract line items list
            items_list = parsed_data.get("line_items", [])

            if not items_list:
                print(f"   ⚠️ No items extracted from {Path(file_path).name}")
                return pd.DataFrame()

            # Create DataFrame
            df = pd.DataFrame(items_list)

            # Add Source Metadata
            df["File Name"] = Path(file_path).name
            df["Extraction Confidence"] = "High" # Placeholder for future logic

            # Normalize Columns for PMS Step
            if "Price" in df.columns:
                df = df.rename(columns={"Price": "Price Per Buying Unit"})

            return df

        except json.JSONDecodeError:
            print(f"   ❌ JSON Parsing Failed for {Path(file_path).name}")
            # In a full production system, we might dump the raw text to a log here
            return pd.DataFrame()
        except Exception as e:
            print(f"   ❌ Unexpected Error processing {Path(file_path).name}: {e}")
            time.sleep(RETRY_DELAY)

    return pd.DataFrame()

def batch_process_invoices(file_paths: List[str], country_code: str) -> pd.DataFrame:
    """
    Orchestrates parallel processing of multiple invoices.
    """
    print(f"\n🚀 Starting Extraction Batch for {len(file_paths)} files...")
    all_extracted_dfs = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Submit all tasks
        future_to_file = {
            executor.submit(process_single_invoice, fp, country_code): fp
            for fp in file_paths
        }

        # Process as they complete
        for future in tqdm(as_completed(future_to_file), total=len(file_paths), desc="Processing Invoices", unit="doc"):
            file_path = future_to_file[future]
            try:
                df_result = future.result()
                if not df_result.empty:
                    all_extracted_dfs.append(df_result)
            except Exception as e:
                print(f"   ☠️ Critical Failure on {Path(file_path).name}: {e}")

    if not all_extracted_dfs:
        print("❌ Batch Extraction Failed: No data retrieved.")
        return pd.DataFrame()

    print("✅ Batch Extraction Complete. Merging data...")
    combined_df = pd.concat(all_extracted_dfs, ignore_index=True)
    return combined_df

# ----------------------------------------------------------------------------
# PART 6: CORE ENGINE - PMS CONVERSION (AI CLEANING & STANDARDIZATION)
# ----------------------------------------------------------------------------

def generate_pms_system_prompt(config: Dict) -> str:
    """Generates the context-aware system prompt for PMS Standardization."""
    country_info = COUNTRY_TAX_DB.get(config['country'], COUNTRY_TAX_DB['AE'])
    translation_mode = "Translate all foreign text to English" if config.get('translate_enabled') else "Keep original language"

    prompt = f"""You are SupyConverter v4.0, a specialized F&B Data Standardization Engine.

CONTEXT:
- Region: {country_info['name']}
- Currency: {country_info['currency']}
- Tax Rate: {country_info['rate']}%
- Translation Mode: {translation_mode}

TASK:
You will receive a list of raw invoice items. You must standardize them into the Supy PMS format.

OUTPUT SCHEMA (JSON Array of Objects):
[
  {{
    "Match %": integer (0-100 confidence),
    "Remarks": "string (CRITICAL/ERROR/WARNING/INFO)",
    "Supplier Item Name": "original string",
    "Supplier Item Code": "original string",
    "Supplier Name": "original string",
    "Buying Unit": "original string",
    "Price Per Buying Unit": number,
    "Base Item / Ingredient Name": "Cleaned Name",
    "Main Category": "FOOD/BEVERAGES/NON FOOD",
    "Sub Category": "string from taxonomy",
    "Base Unit (Kg / L / Piece)": "Kg/L/Piece",
    "Qty in Base Packaging": number,
    "Package Name": "Bottle/Can/Box/Bag/Case/etc or empty",
    "Base Package Multiplier": number or null,
    "Larger Package Name": "string or empty",
    "Bigger Packaging": "string or empty",
    "Par Level": null,
    "Min Level": null,
    "Is Item Taxable?": "Yes/No",
    "Base Item Prep Wastage (%)": null,
    "Affects COGS (Yes/No)": "Yes"
  }}
]

STRICT STANDARDIZATION RULES:

1. **Base Item Name Cleaning**:
   - For FOOD: Remove brands, origins, grades, and packaging sizes. Keep it generic (3-5 words).
     - Example: "Fresh USA Strawberry Grade A 250g" -> "Strawberry Fresh"
     - Example: "Kikkoman Soy Sauce 1L" -> "Soy Sauce"
   - For ALCOHOL: Keep Brand, Vintage, and Type.
     - Example: "Heineken Beer" -> "Heineken Beer"
     - Example: "Chateau Margaux 2015" -> "Chateau Margaux 2015"

2. **Unit Standardization**:
   - 'Base Unit' must ONLY be: "Kg" (solids), "L" (liquids), or "Piece" (everything else).
   - Convert grams to Kg (500g -> 0.5 Kg).
   - Convert ml to L (750ml -> 0.75 L).

3. **Package Name**:
   - Must be a physical container type (Bottle, Can, Jar, Tin, Bag, Box, Case, Tub).
   - NEVER use a unit of measure (Kg, L, ml) as a Package Name.

4. **Taxability**:
   - Check if item is Food. If country is {country_info['name']} ({country_info['food_exempt']}), set "Is Item Taxable?" to "No" for food.
   - Alcohol and Non-Food are always "Yes".

5. **Remarks & Scoring**:
   - If information is missing (e.g., no price), set "Remarks" to "CRITICAL: Missing Price".
   - If unit is ambiguous, set "Remarks" to "WARNING: Check Unit".
   - "Match %": Start at 100. Deduct 10 for warnings, 30 for critical errors.

Process the input array and return the JSON array.
"""
    return prompt

def process_pms_batch_chunk(items_batch: List[Dict], config: Dict) -> List[Dict]:
    """Processes a small batch of items through the AI model."""
    system_prompt = generate_pms_system_prompt(config)
    safe_data = make_serializable(items_batch)

    api_payload = {
        "model": MODEL_ID,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps({"items": safe_data})}
        ],
        "temperature": 0.1,
        "max_tokens": OUTPUT_MAX_TOKENS
    }

    headers = {"Authorization": f"Bearer {OPENROUTER_API_KEY}"}

    for attempt in range(RETRY_ATTEMPTS):
        try:
            response = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=api_payload, timeout=120)

            if response.status_code == 200:
                content = response.json()["choices"][0]["message"]["content"]
                content = re.sub(r"```json|```", "", content).strip()
                result_json = json.loads(content)

                # Validation: Ensure output count matches input count
                if len(result_json) != len(items_batch):
                    # Pad results if mismatch occurs
                    padding = len(items_batch) - len(result_json)
                    for _ in range(padding):
                        result_json.append({
                            "Supplier Item Name": "ERROR",
                            "Remarks": "API Response Count Mismatch - Failed to Process"
                        })
                return result_json
            else:
                time.sleep(2)
        except Exception as e:
            print(f"   ⚠️ Batch Error: {e}")
            time.sleep(2)

    # Fallback if all retries fail
    fallback_results = []
    for item in items_batch:
        fallback_item = item.copy()
        fallback_item["Remarks"] = "CRITICAL: API Processing Failed"
        fallback_results.append(fallback_item)
    return fallback_results

def run_pms_conversion_engine(df: pd.DataFrame, config: Dict) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Main controller for the PMS conversion workflow.
    Splits data into batches, processes them, and performs post-processing analysis.
    """
    print(f"\n🧠 Starting PMS AI Engine for {len(df)} items...")

    # 1. Pre-Processing: Column Normalization
    df_clean = df.copy()

    # Map common column names to "Supplier Item Name"
    name_candidates = ["Description", "Item", "Product", "Product Name", "Item Name", "Material"]
    if "Supplier Item Name" not in df_clean.columns:
        for candidate in name_candidates:
            if candidate in df_clean.columns:
                df_clean = df_clean.rename(columns={candidate: "Supplier Item Name"})
                break

    # Ensure Price column exists
    if "Price Per Buying Unit" not in df_clean.columns:
        price_candidates = ["Price", "Unit Price", "Cost", "Amount"]
        for candidate in price_candidates:
            if candidate in df_clean.columns:
                df_clean = df_clean.rename(columns={candidate: "Price Per Buying Unit"})
                break

    # Convert to list of dicts for batching
    all_records = df_clean.to_dict('records')
    batches = [all_records[i:i + BATCH_SIZE] for i in range(0, len(all_records), BATCH_SIZE)]

    processed_results = []

    # 2. Parallel Processing
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_pms_batch_chunk, batch, config) for batch in batches]

        for future in tqdm(as_completed(futures), total=len(batches), desc="Standardizing Data"):
            try:
                processed_results.extend(future.result())
            except Exception as e:
                print(f"   ❌ Batch Worker Failed: {e}")

    # 3. Create Main DataFrame
    pms_df = pd.DataFrame(processed_results)

    # 4. Post-Processing: Duplicate Detection (Fuzzy Logic)
    print("   🔍 analyzing for duplicates...")
    duplicates = []
    if 'Base Item / Ingredient Name' in pms_df.columns:
        base_names = pms_df['Base Item / Ingredient Name'].astype(str).tolist()

        # Optimization: Only run N^2 check if dataset is manageable
        if len(base_names) < 3000:
            for i in range(len(base_names)):
                for j in range(i + 1, len(base_names)):
                    # Check similarity
                    ratio = fuzz.ratio(base_names[i].lower(), base_names[j].lower())
                    if ratio > 90:
                        duplicates.append({
                            'Original Row A': i + 2,
                            'Item A': base_names[i],
                            'Original Row B': j + 2,
                            'Item B': base_names[j],
                            'Similarity Score': ratio
                        })

    duplicates_df = pd.DataFrame(duplicates)

    # 5. Post-Processing: Review Queue Generation
    # Filter for items that need human attention
    print("   🚩 generating review queue...")

    # Filter conditions
    has_critical = pms_df['Remarks'].astype(str).str.contains('CRITICAL', case=False, na=False)
    has_error = pms_df['Remarks'].astype(str).str.contains('ERROR', case=False, na=False)
    has_warning = pms_df['Remarks'].astype(str).str.contains('WARNING', case=False, na=False)
    low_confidence = pd.to_numeric(pms_df.get('Match %', 100), errors='coerce') < 80

    review_mask = has_critical | has_error | has_warning | low_confidence
    review_df = pms_df[review_mask].copy()

    # Add reference to original row number
    if not review_df.empty:
        review_df.insert(0, "Row Number", review_df.index + 2)

    return pms_df, review_df, duplicates_df

# ----------------------------------------------------------------------------
# PART 7: FINAL OUTPUT GENERATION (EXCEL FORMATTING)
# ----------------------------------------------------------------------------

def create_formatted_excel(raw_df: pd.DataFrame, pms_df: pd.DataFrame, review_df: pd.DataFrame, dup_df: pd.DataFrame, config: Dict) -> str:
    """
    Generates the professional 5-sheet Excel workbook with conditional formatting.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Supy_Unified_Export_{config['country']}_{timestamp}.xlsx"

    print(f"\n💾 Writing Output File: {filename}")

    # 1. Write Data
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1: RAW_DATA
        raw_df.to_excel(writer, sheet_name='RAW_DATA', index=False)

        # Sheet 2: PMS_Comparison (The Main Sheet)
        pms_df.to_excel(writer, sheet_name='PMS_Comparison', index=False)

        # Sheet 3: Review_Queue
        if not review_df.empty:
            review_df.to_excel(writer, sheet_name='Review_Queue', index=False)
        else:
            pd.DataFrame({"Message": ["No items flagged for review! Great job."]}).to_excel(writer, sheet_name='Review_Queue', index=False)

        # Sheet 4: Potential_Duplicates
        if not dup_df.empty:
            dup_df.to_excel(writer, sheet_name='Potential_Duplicates', index=False)
        else:
            pd.DataFrame({"Message": ["No duplicates detected."]}).to_excel(writer, sheet_name='Potential_Duplicates', index=False)

        # Sheet 5: Summary_Stats
        stats_data = {
            "Metric": ["Processing Date", "Target Country", "Total Items Processed", "Review Queue Count", "Duplicate Pairs", "Translation Enabled"],
            "Value": [datetime.now().strftime('%Y-%m-%d %H:%M'), config['country'], len(pms_df), len(review_df), len(dup_df), str(config['translate_enabled'])]
        }
        pd.DataFrame(stats_data).to_excel(writer, sheet_name='Summary', index=False)

    # 2. Apply Visual Styling (OpenPyXL)
    print("   🎨 Applying Conditional Formatting & Styles...")
    wb = load_workbook(filename)

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Supy Blue

    fill_critical = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    fill_warning = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid") # Light Orange

    border_style = Side(style='thin', color='000000')
    full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Style PMS_Comparison Sheet
    ws = wb['PMS_Comparison']

    # Identify key column indices
    headers = [cell.value for cell in ws[1]]
    try:
        idx_match = headers.index("Match %")
        idx_remarks = headers.index("Remarks")
    except ValueError:
        idx_match = -1
        idx_remarks = -1

    # Loop through rows to style
    for row in ws.iter_rows():
        for cell in row:
            # Header Styling
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = full_border
            else:
                # Content Styling
                cell.border = full_border

                # Conditional Logic
                if idx_match != -1 and idx_remarks != -1:
                    match_val = row[idx_match].value
                    remarks_val = str(row[idx_remarks].value).upper()

                    # Red for Critical/Error
                    if "CRITICAL" in remarks_val or "ERROR" in remarks_val:
                        cell.fill = fill_critical
                    # Orange for Warnings or Low Match
                    elif "WARNING" in remarks_val or (isinstance(match_val, (int, float)) and match_val < 80):
                        cell.fill = fill_warning

    # Auto-Fit Column Widths for all sheets
    for sheet_name in wb.sheetnames:
        ws_curr = wb[sheet_name]
        for column in ws_curr.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60) # Cap width at 60
            ws_curr.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    return filename

# ----------------------------------------------------------------------------
# PART 8: MAIN EXECUTION CONTROLLER
# ----------------------------------------------------------------------------

def main():
    # Header UI
    print("\n" + "="*80)
    print("   🥗 SUPY VISION CONVERTER v4.0 (UNABRIDGED ENTERPRISE EDITION)")
    print("   ------------------------------------------------------------")
    print("   Engine A: Invoice Extraction (OCR)")
    print("   Engine B: PMS Standardization (AI)")
    print("   Engine C: Global Tax Compliance (200+ Countries)")
    print("="*80 + "\n")

    # 1. Initialize Drive (if applicable)
    drive_service = None
    if COLAB_ENV:
        print("📥 INIT: Checking Environment...")
        try:
            drive_service = authenticate_google_drive()
        except:
            print("   ⚠️ Drive Authentication Skipped.")

    # 2. User Configuration
    print("\n🌍 STEP 1: REGIONAL CONFIGURATION")
    print("   Available Presets: AE (UAE), SA (Saudi), IN (India), US (USA), UK (Great Britain), etc.")
    user_country = input("   👉 Enter 2-Letter Country Code [Default: AE]: ").strip().upper()

    if not user_country:
        user_country = "AE"

    if user_country in COUNTRY_TAX_DB:
        c_data = COUNTRY_TAX_DB[user_country]
        print(f"   ✅ Context Set: {c_data['name']} (Tax: {c_data['rate']}%)")
    else:
        print(f"   ⚠️ Unknown Code '{user_country}'. Defaulting to UAE (AE).")
        user_country = "AE"

    config = {
        "country": user_country,
        "translate_enabled": False, # Will be set later
        "verify_prices": True
    }

    # 3. Input Source Selection
    print("\n📂 STEP 2: DATA INGESTION")
    print("   1. Direct File Upload (Computer)")
    print("   2. Google Drive Folder Import")

    source_choice = input("   👉 Select Source [1]: ").strip()

    all_files_to_process = []
    temp_workspace = f"workspace_{int(time.time())}"
    os.makedirs(temp_workspace, exist_ok=True)

    if source_choice == "2" and drive_service:
        folder_url = input("   👉 Paste Google Drive Folder Link/ID: ").strip()
        folder_id = extract_drive_folder_id(folder_url)

        print(f"   🔍 Scanning Drive Folder: {folder_id}...")
        try:
            results = drive_service.files().list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="files(id, name)"
            ).execute()
            items = results.get('files', [])

            if not items:
                print("   ❌ No files found in that folder.")
                return

            print(f"   📥 Downloading {len(items)} files...")
            for item in tqdm(items, desc="Downloading"):
                file_id = item['id']
                file_name = item['name']
                local_path = os.path.join(temp_workspace, file_name)

                request = drive_service.files().get_media(fileId=file_id)
                with io.FileIO(local_path, 'wb') as fh:
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while not done:
                        _, done = downloader.next_chunk()
                all_files_to_process.append(local_path)

        except Exception as e:
            print(f"   ❌ Drive Error: {e}")
            return
    else:
        print("   📤 Please upload your files via the widget below...")
        try:
            uploaded = files.upload()
            for filename in uploaded.keys():
                # Move to workspace for cleanliness
                new_path = os.path.join(temp_workspace, filename)
                os.rename(filename, new_path)
                all_files_to_process.append(new_path)
        except:
            # Fallback for local python execution (non-colab)
            local_input = input("   👉 Enter local file path: ").strip()
            if os.path.exists(local_input):
                all_files_to_process.append(local_input)

    # 4. Recursive Processing (Unzip)
    print("\n📦 STEP 3: FILE PREPARATION")
    final_file_list = []

    for file_path in all_files_to_process:
        if file_path.lower().endswith(".zip"):
            extracted = recursive_zip_extractor(file_path, os.path.join(temp_workspace, "unzipped"))
            final_file_list.extend(extracted)
        else:
            final_file_list.append(file_path)

    unique_files = list(set(final_file_list))
    print(f"   ✅ Ready to process {len(unique_files)} unique files.")

    # 5. Workflow Determination
    has_images = any(f.lower().endswith(tuple(INVOICE_EXTENSIONS)) for f in unique_files)
    has_excel = any(f.lower().endswith(tuple(EXCEL_EXTENSIONS)) for f in unique_files)

    print("\n⚙️ STEP 4: SELECT MODE")
    if has_images:
        print("   [1] Full Pipeline: Invoice Extraction -> PMS Conversion")
        print("   [2] Extraction Only: Invoice -> Raw Excel")
    if has_excel:
        print("   [3] Conversion Only: Raw Excel -> PMS Excel")

    mode_choice = input("   👉 Select Mode [1]: ").strip() or "1"

    # 6. Execution Phase
    raw_dataframe = pd.DataFrame()

    # --- PHASE A: EXTRACTION ---
    if mode_choice in ["1", "2"] and has_images:
        image_files = [f for f in unique_files if f.lower().endswith(tuple(INVOICE_EXTENSIONS))]
        if image_files:
            raw_dataframe = batch_process_invoices(image_files, config['country'])

            if mode_choice == "2":
                # Save and Exit
                output_name = f"Supy_Raw_Extraction_{int(time.time())}.xlsx"
                raw_dataframe.to_excel(output_name, index=False)
                print(f"\n✅ DONE. Download: {output_name}")
                if COLAB_ENV: files.download(output_name)
                return

    # --- PHASE B: DATA LOADING (If Mode 3) ---
    elif mode_choice == "3" and has_excel:
        excel_files = [f for f in unique_files if f.lower().endswith(tuple(EXCEL_EXTENSIONS))]
        if excel_files:
            print(f"   📖 Loading Data from: {Path(excel_files[0]).name}")
            raw_dataframe = pd.read_excel(excel_files[0])

            # Language Detection Check
            try:
                sample_texts = raw_dataframe.iloc[:15, 0].astype(str).tolist() # Check first column
                detected_lang = Counter([detect_text_language(t) for t in sample_texts]).most_common(1)[0][0]

                if detected_lang != "English":
                    print(f"   🗣️ Detected Language: {detected_lang}")
                    trans_input = input("   👉 Enable AI Translation to English? (y/n) [n]: ").lower()
                    if trans_input == 'y':
                        config['translate_enabled'] = True
                        print("   🌍 Translation Mode: ENABLED")
            except:
                pass

    # --- PHASE C: PMS CONVERSION ---
    if not raw_dataframe.empty:
        pms_df, review_df, dup_df = run_pms_conversion_engine(raw_dataframe, config)

        # --- PHASE D: EXPORT ---
        final_file = create_formatted_excel(raw_dataframe, pms_df, review_df, dup_df, config)

        print("\n" + "="*80)
        print("🎉 PROCESSING COMPLETE SUCCESSFULLY")
        print(f"   📄 Final Report: {final_file}")
        print("="*80)

        if COLAB_ENV:
            print("   ⬇️ Initiating Download...")
            files.download(final_file)
    else:
        print("\n❌ Error: No data found to process.")


if __name__ == "__main__":
    # main()  <-- THIS IS THE ONLY CHANGE. Prevents CLI from running in Streamlit.
    pass